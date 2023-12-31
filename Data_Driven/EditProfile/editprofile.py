# Generated by Selenium IDE
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

from selenium.common.exceptions import NoSuchElementException

class FileExcelReader:
    file = ""
    sheetName = ""

    def __init__(self, file, sheetName):
        self.file = file
        self.sheetName = sheetName

    def getRowCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_row)

    def getColumnCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_column)

    def readData(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    def writeData(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)

class TestEditProfile():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.maximize_window()
    self.vars = {}

    self.driver.get('https://sandbox.moodledemo.net/login/index.php')
    usernameInput = self.driver.find_element(By.NAME,"username")
    passwordInput = self.driver.find_element(By.NAME,"password")
    submitBtn = self.driver.find_element(By.ID,"loginbtn")
    usernameInput.clear()
    usernameInput.send_keys("teacher")
    passwordInput.send_keys("sandbox")    
    submitBtn.click()
    time.sleep(1)
  
  def teardown_method(self):
    self.driver.quit()

  def test_edit_profile(self, _fname, _lname, expectedResult):
    
    avatarIcon = self.driver.find_element(By.XPATH,'//*[@id="user-menu-toggle"]')
    avatarIcon.click()
    time.sleep(1)
    selectProfile = self.driver.find_element(By.XPATH,'//*[@id="carousel-item-main"]/a[1]')
    selectProfile.click()
    time.sleep(1)

    selectEditProfile = self.driver.find_element(By.XPATH,'//*[@id="region-main"]/div/div/div/section[1]/div/ul/li[1]/span/a')
    selectEditProfile.click()
    time.sleep(1)

    fname = self.driver.find_element(By.XPATH,'//*[@id="id_firstname"]')
    fname.clear()
    if _fname: fname.send_keys(_fname)
    time.sleep(1)

    lname = self.driver.find_element(By.XPATH,'//*[@id="id_lastname"]')
    lname.clear()
    if _lname: lname.send_keys(_lname) 
    time.sleep(1)

    updateBtn = self.driver.find_element(By.XPATH,'//*[@id="id_submitbutton"]')
    updateBtn.click()
    time.sleep(1)
    

    if expectedResult == "Success":
        alert = self.driver.find_element(By.XPATH,'/html/body/div[2]/div[3]/div/div[2]/div/section/span/div')
        return alert.text == "Changes saved"
    elif expectedResult == "Not success":
        return self.driver.current_url == 'https://sandbox.moodledemo.net/user/edit.php'
    
    return True
    # if expectedResult == "Success": #done
    #     assert self.driver.current_url == 'https://sandbox.moodledemo.net/'
    #     profileTag = self.driver.find_element(By.XPATH,'//*[@id="user-menu-toggle"]')
    #     profileTag.click()
    #     time.sleep(1)
    #     logoutBtn = self.driver.find_element(By.XPATH,'//*[@id="carousel-item-main"]/a[9]')
    #     logoutBtn.click()
    #     time.sleep(1)
    #     loginPageBtn = self.driver.find_element(By.XPATH,'//*[@id="usernavigation"]/div[5]/div/span/a')
    #     loginPageBtn.click()
    #     time.sleep(2)
    #     return True
    # elif expectedResult == "Not success": #done
    #     errorNotification = self.driver.find_element(By.XPATH,'//*[@id="region-main"]/div/div/div/div/div[1]')
    #     return True


if __name__ == "__main__":
    excel = FileExcelReader('SecB_editprofile_data.xlsx', 'Sheet1')
    test = TestEditProfile()
    test.setup_method()
    nRows = excel.getRowCount()
    for row in range(2, nRows + 1):
        _fname = excel.readData(row,1)
        _lname = excel.readData(row,2)
        expectedResult = excel.readData(row,3)
        try:
            result = test.test_edit_profile( _fname, _lname, expectedResult)
            excel.writeData("Passed",row,4)
        except:
            excel.writeData("Failed",row,4)


    test.teardown_method()