# Generated by Selenium IDE
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

from selenium.common.exceptions import NoSuchElementException

class FileExcelReader:
    file = ""
    sheetName = ""

    def __init__(self, file, sheetName):
        self.file = file
        self.sheetName = sheetName

    def getRowCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_row)

    def getColumnCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_column)

    def readData(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    def writeData(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)

class TestSendMessage():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.maximize_window()
    self.vars = {}

    self.driver.get('https://sandbox.moodledemo.net/login/index.php')
    usernameInput = self.driver.find_element(By.NAME,"username")
    passwordInput = self.driver.find_element(By.NAME,"password")
    submitBtn = self.driver.find_element(By.ID,"loginbtn")
    usernameInput.clear()
    usernameInput.send_keys("admin")
    passwordInput.send_keys("sandbox")    
    submitBtn.click()
    time.sleep(1)
  
  def teardown_method(self):
    self.driver.quit()

  def test_send_message(self, text, expectedResult):
    messageIcon = self.driver.find_element(By.CSS_SELECTOR,'#usernavigation > div:nth-child(4)')
    messageIcon.click()

    time.sleep(4)
    selectUser = self.driver.find_element(By.CSS_SELECTOR,'.px-2.py-1.small.position-absolute.position-right')
    selectUser.click()
    time.sleep(4)

    inputMessage = self.driver.find_element(By.CSS_SELECTOR,'.p-sm-2>.d-flex.mt-sm-1>textarea')
    if text: inputMessage.send_keys(text)
    time.sleep(1)
    
    textContainer = self.driver.find_element(By.CSS_SELECTOR,'*>div[data-region="day-messages-container"]')
    prevChildren = textContainer.get_property("childElementCount")

    buttonSend = self.driver.find_element(By.XPATH,'/html/body/div[2]/div[4]/div/div[4]/div[1]/div[1]/div[2]/div/button[2]')
    buttonSend.click()
    time.sleep(3)
    
    if expectedResult == "Success": #done
        # assert self.driver.current_url == 'https://sandbox.moodledemo.net/'
        assert prevChildren + 1 == textContainer.get_property("childElementCount") 
        time.sleep(1)

        return True
    elif expectedResult == "Not success": #done
        assert prevChildren == textContainer.get_property("childElementCount") 
        time.sleep(1)
        
        return True
    messageIcon.click()


if __name__ == "__main__":
    excel = FileExcelReader('SecB_sendmessage_data.xlsx', 'Sheet1')
    test = TestSendMessage()
    # test.setup_method()
    nRows = excel.getRowCount()
    for row in range(2, nRows + 1):
        text = excel.readData(row,1)
        expectedResult = excel.readData(row,2)
        test.setup_method()
        try:
            result = test.test_send_message(text, expectedResult)
            excel.writeData("Passed",row,3)
        except:
            excel.writeData("Failed",row,3)
        test.teardown_method()


    # test.teardown_method()