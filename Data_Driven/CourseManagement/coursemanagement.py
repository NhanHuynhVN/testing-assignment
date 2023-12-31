# Generated by Selenium IDE
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

from selenium.common.exceptions import NoSuchElementException

class FileExcelReader:
    file = ""
    sheetName = ""

    def __init__(self, file, sheetName):
        self.file = file
        self.sheetName = sheetName

    def getRowCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_row)

    def getColumnCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_column)

    def readData(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    def writeData(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)

class TestCourseManagementProfile():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.maximize_window()
    self.vars = {}

    self.driver.get('https://sandbox.moodledemo.net/login/index.php')
    usernameInput = self.driver.find_element(By.NAME,"username")
    passwordInput = self.driver.find_element(By.NAME,"password")
    submitBtn = self.driver.find_element(By.ID,"loginbtn")
    usernameInput.clear()
    usernameInput.send_keys("admin")
    passwordInput.send_keys("sandbox")    
    submitBtn.click()
    time.sleep(1)
  
  def teardown_method(self):
    self.driver.quit()

  def test_create_course(self ,_fullName, _shortName,_category, expectedResult):
    
    myCourses = self.driver.find_elements(By.CLASS_NAME,'nav-item')
    myCourses[2].click()
    buttons = self.driver.find_elements(By.CLASS_NAME,"singlebutton")
    buttons[1].click()
    # myCourse.find_element(By.CSS_SELECTOR,">a").click()
    time.sleep(5)

    fullname = self.driver.find_element(By.XPATH,'//*[@id="id_fullname"]')
    fullname.clear()
    if _fullName: fullname.send_keys(_fullName)
    time.sleep(1)

    shortName = self.driver.find_element(By.XPATH,'//*[@id="id_shortname"]')
    shortName.clear()
    if _shortName: shortName.send_keys(_shortName)
    time.sleep(1)
    saveBtn = self.driver.find_element(By.XPATH,'//*[@id="id_saveanddisplay"]')
    saveBtn.click()
    time.sleep(1)
    
    if expectedResult == "Success":
        return self.driver.current_url == 'https://sandbox.moodledemo.net/course/view.php?id=6'
    elif expectedResult == "Not success":
        return self.driver.current_url == 'https://sandbox.moodledemo.net/course/edit.php'
    return True


if __name__ == "__main__":
    excel = FileExcelReader('SecB_coursemanagement_data.xlsx', 'Sheet1')
    test = TestCourseManagementProfile()
    test.setup_method()

    nRows = excel.getRowCount()
    for row in range(2, nRows + 1):
        _fullName = excel.readData(row,1)
        _shortName = excel.readData(row,2)
        _category = excel.readData(row,3)
        expectedResult = excel.readData(row,4)
        try:
            result = test.test_create_course( _fullName, _shortName,_category, expectedResult)
            excel.writeData("Passed",row,5)
        except:
            excel.writeData("Failed",row,5)


    test.teardown_method()